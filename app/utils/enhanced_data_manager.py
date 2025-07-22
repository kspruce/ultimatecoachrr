#!/usr/bin/env python3
"""
Enhanced Data Manager for Flask Application

Features:
1. Export database to Excel files for easy viewing
2. Create downloadable ZIP files for local storage
3. Import data from local ZIP files
4. View detailed model data in the browser
"""

from app import db
import os
import json
import shutil
import zipfile
import tempfile
import pandas as pd
import logging
from datetime import datetime
from sqlalchemy import inspect
import importlib
import pkgutil
from flask import send_file
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename
import math

# Configure logging
logging.basicConfig(level=logging.INFO, 
                    format='[%(asctime)s] %(levelname)s: %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

class EnhancedDataManager:
    """Enhanced data manager with Excel export and ZIP functionality"""
    
    def __init__(self, export_dir=None):
        """Initialize the data manager"""
        # Always default to /tmp if no directory is specified
        self.export_dir = export_dir if export_dir else '/tmp/data_exports'
        
        # Create export directory if it doesn't exist
        os.makedirs(self.export_dir, exist_ok=True)
        
        # Load all models
        self.models = self._load_models()
        
        # Determine dependency order for import/export
        self.dependency_order = self._determine_dependency_order()
    
    def _load_models(self):
        """Dynamically load all SQLAlchemy models"""
        models = {}
        
        # Get the app.models package
        import app.models
        
        # Walk through all modules in the models package
        for _, name, is_pkg in pkgutil.iter_modules(app.models.__path__, app.models.__name__ + '.'):
            if not is_pkg:
                try:
                    # Import the module
                    module = importlib.import_module(name)
                    
                    # Find all SQLAlchemy models in the module
                    for attr_name in dir(module):
                        attr = getattr(module, attr_name)
                        if isinstance(attr, type) and hasattr(attr, '__tablename__'):
                            # This is a SQLAlchemy model
                            models[attr.__tablename__] = attr
                except ImportError as e:
                    logger.warning(f"Could not import module {name}: {e}")
        
        return models
    
    def _determine_dependency_order(self):
        """Determine the order in which tables should be processed based on foreign key dependencies"""
        dependency_graph = {}
        
        # Build dependency graph
        for table_name, model in self.models.items():
            dependencies = set()
            for column in model.__table__.columns:
                for foreign_key in column.foreign_keys:
                    target_table = foreign_key.column.table.name
                    if target_table != table_name:  # Avoid self-references
                        dependencies.add(target_table)
            dependency_graph[table_name] = dependencies
        
        # Topological sort
        visited = set()
        temp_visited = set()
        order = []
        
        def visit(table):
            if table in temp_visited:
                # Cyclic dependency, handle gracefully
                return
            if table in visited:
                return
            
            temp_visited.add(table)
            
            # Visit dependencies first
            for dep in dependency_graph.get(table, set()):
                if dep in self.models:  # Only consider dependencies that are actual models
                    visit(dep)
            
            temp_visited.remove(table)
            visited.add(table)
            order.append(table)
        
        # Visit all tables
        for table in dependency_graph:
            if table not in visited:
                visit(table)
        
        # Reverse to get correct order (dependencies first)
        return list(reversed(order))
        
    def get_model_info(self):
        """Get information about all models"""
        model_info = {
            'models': {},
            'total_models': len(self.models),
            'dependency_order': self.dependency_order
        }
        
        for table_name, model in self.models.items():
            # Get column information
            columns = []
            foreign_keys = []
            
            for column in model.__table__.columns:
                column_info = {
                    'name': column.name,
                    'type': str(column.type),
                    'nullable': column.nullable,
                    'primary_key': column.primary_key
                }
                columns.append(column_info)
                
                # Check for foreign keys
                for fk in column.foreign_keys:
                    fk_info = {
                        'column': column.name,
                        'references': f"{fk.column.table.name}.{fk.column.name}"
                    }
                    foreign_keys.append(fk_info)
            
            # Count records
            try:
                record_count = model.query.count()
            except Exception as e:
                logger.warning(f"Could not count records for {table_name}: {e}")
                record_count = 0
            
            model_info['models'][table_name] = {
                'class_name': model.__name__,
                'columns': columns,
                'foreign_keys': foreign_keys,
                'record_count': record_count
            }
        
        return model_info
    
    def export_all_data(self, timestamp=True, custom_name=None, format='json', in_memory=False):
        """Export all data to in-memory ZIP file"""
        if custom_name:
            export_dir_name = custom_name
        elif timestamp:
            export_dir_name = f"data_exports_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        else:
            export_dir_name = "data_exports"
        
        # Create in-memory ZIP file
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add metadata
            metadata = {
                'export_timestamp': datetime.now().isoformat(),
                'total_models': len(self.models),
                'format': format
            }
            zipf.writestr('metadata.json', json.dumps(metadata, indent=2))
            
            # Export summary
            export_summary = {
                'status': 'in_progress',
                'timestamp': datetime.now().isoformat(),
                'format': format
            }
            
            # Process each model in dependency order
            for table_name in self.dependency_order:
                if table_name in self.models:
                    model = self.models[table_name]
                    
                    try:
                        # Get all records
                        records = model.query.all()
                        
                        # Convert to list of dictionaries
                        data = []
                        for record in records:
                            record_dict = {}
                            for column in model.__table__.columns:
                                value = getattr(record, column.name)
                                # Handle datetime objects
                                if isinstance(value, datetime):
                                    value = value.isoformat()
                                record_dict[column.name] = value
                            data.append(record_dict)
                        
                        # Add JSON to ZIP
                        zipf.writestr(f"{table_name}.json", json.dumps(data, indent=2))
                        
                        # If Excel format is requested, also add Excel
                        if format == 'excel' or format == 'both':
                            excel_buffer = io.BytesIO()
                            self._save_as_excel(data, excel_buffer, table_name)
                            excel_buffer.seek(0)
                            zipf.writestr(f"{table_name}.xlsx", excel_buffer.getvalue())
                        
                        # Update summary
                        export_summary[table_name] = {
                            'status': 'completed',
                            'records_exported': len(data)
                        }
                        
                        logger.info(f"Exported {len(data)} records from {table_name}")
                        
                    except Exception as e:
                        logger.error(f"Error exporting {table_name}: {e}")
                        export_summary[table_name] = {
                            'status': 'error',
                            'error': str(e)
                        }
            
            # Update and save summary
            export_summary['status'] = 'completed'
            export_summary['completed_timestamp'] = datetime.now().isoformat()
            zipf.writestr('export_summary.json', json.dumps(export_summary, indent=2))
        
        memory_file.seek(0)
        return memory_file, f"{export_dir_name}.zip"


    
    def _save_as_excel(self, data, file_path_or_object, table_name):
        """
        Save data as an Excel file with formatting
        
        Parameters:
        - data: List of dictionaries containing the data
        - file_path_or_object: Either a string file path or a file-like object (BytesIO)
        - table_name: Name of the table/sheet
        """
        if not data:
            # Create empty DataFrame with column names
            df = pd.DataFrame()
        else:
            # Create DataFrame from data
            df = pd.DataFrame(data)
        
        # Save to Excel - pandas handles both file paths and file-like objects
        with pd.ExcelWriter(file_path_or_object, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=table_name, index=False)
            
            # Get the worksheet
            worksheet = writer.sheets[table_name]
            
            # Format header row
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            # Add table borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(data) + 1):
                for cell in row:
                    cell.border = thin_border

    
    def _create_zip_archive(self, export_path):
        """Create a ZIP archive of the export directory"""
        zip_filename = os.path.basename(export_path) + '.zip'
        zip_path = os.path.join(os.path.dirname(export_path), zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(export_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, export_path)
                    zipf.write(file_path, arcname)
        
        logger.info(f"Created ZIP archive: {zip_path}")
        return zip_path
    
    def import_from_zip(self, zip_file_path, clear_existing=False):
        """Import data from a ZIP file"""
        # Create a temporary directory
        temp_dir = tempfile.mkdtemp()
        logger.info(f"Created temporary directory: {temp_dir}")
        
        try:
            # Check if the file exists and is a ZIP file
            if not os.path.exists(zip_file_path):
                raise ValueError(f"ZIP file does not exist: {zip_file_path}")
            
            if not zipfile.is_zipfile(zip_file_path):
                raise ValueError(f"File is not a valid ZIP file: {zip_file_path}")
            
            logger.info(f"Extracting ZIP file: {zip_file_path}")
            # Extract ZIP file
            with zipfile.ZipFile(zip_file_path, 'r') as zipf:
                # Log the contents of the ZIP file
                file_list = zipf.namelist()
                logger.info(f"ZIP file contents: {file_list}")
                zipf.extractall(temp_dir)
            
            # List all files in the temp directory for debugging
            all_files = []
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    all_files.append(os.path.join(root, file))
            logger.info(f"All extracted files: {all_files}")
            
            # Find metadata.json to determine the root directory
            metadata_path = None
            
            # First, check if metadata.json is at the root level
            root_metadata = os.path.join(temp_dir, 'metadata.json')
            if os.path.exists(root_metadata):
                metadata_path = root_metadata
                logger.info(f"Found metadata.json at root level: {metadata_path}")
            else:
                # Search for metadata.json in subdirectories
                for root, dirs, files in os.walk(temp_dir):
                    if 'metadata.json' in files:
                        metadata_path = os.path.join(root, 'metadata.json')
                        logger.info(f"Found metadata.json in subdirectory: {metadata_path}")
                        break
            
            # If no metadata.json is found, look for any JSON files
            if not metadata_path:
                logger.warning("No metadata.json found, looking for any JSON files")
                json_files = []
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.endswith('.json'):
                            json_files.append(os.path.join(root, file))
                
                if json_files:
                    # Use the directory containing the first JSON file
                    import_dir = os.path.dirname(json_files[0])
                    logger.info(f"Using directory with JSON files: {import_dir}")
                    
                    # Create a simple metadata.json
                    metadata_path = os.path.join(import_dir, 'metadata.json')
                    with open(metadata_path, 'w') as f:
                        json.dump({
                            'export_timestamp': datetime.now().isoformat(),
                            'total_models': len(json_files),
                            'format': 'json'
                        }, f)
                else:
                    raise ValueError("No JSON files found in the ZIP file")
            
            # The directory containing metadata.json is the export root
            import_dir = os.path.dirname(metadata_path)
            logger.info(f"Import directory determined as: {import_dir}")
            
            # Import data from the extracted directory
            return self.import_all_data(import_dir, clear_existing)
            
        except Exception as e:
            logger.error(f"Error in import_from_zip: {e}", exc_info=True)
            raise
        finally:
            # Clean up temporary directory
            logger.info(f"Cleaning up temporary directory: {temp_dir}")
            shutil.rmtree(temp_dir)


    
    def import_all_data(self, import_dir, clear_existing=False):
        """Import all data from JSON files"""
        logger.info(f"Starting import from directory: {import_dir}")
        
        if not os.path.exists(import_dir):
            logger.error(f"Import directory does not exist: {import_dir}")
            raise ValueError(f"Import directory does not exist: {import_dir}")
        
        # Check for metadata.json
        metadata_path = os.path.join(import_dir, 'metadata.json')
        if not os.path.exists(metadata_path):
            logger.error(f"metadata.json not found in {import_dir}")
            raise ValueError(f"Invalid export directory: metadata.json not found in {import_dir}")
        
        # Load metadata
        with open(metadata_path, 'r') as f:
            metadata = json.load(f)
            logger.info(f"Loaded metadata: {metadata}")
        
        # Import summary
        import_summary = {
            'status': 'in_progress',
            'timestamp': datetime.now().isoformat(),
            'metadata': metadata,
            'results': {}
        }
        
        # Process each model in dependency order
        # If clear_existing is True, delete all data in reverse dependency order
        if clear_existing:
            logger.info("Clearing existing data before import")
            for table_name in reversed(self.dependency_order):
                if table_name in self.models:
                    try:
                        model = self.models[table_name]
                        db.session.query(model).delete()
                        db.session.commit()
                        logger.info(f"Cleared all data from {table_name}")
                    except Exception as e:
                        logger.error(f"Error clearing {table_name}: {e}")
                        db.session.rollback()
                        import_summary['results'][table_name] = {
                            'status': 'error',
                            'error': f"Failed to clear table: {str(e)}"
                        }
        
        # Import data in dependency order
        logger.info("Starting data import in dependency order")
        for table_name in self.dependency_order:
            if table_name in self.models:
                model = self.models[table_name]
                json_path = os.path.join(import_dir, f"{table_name}.json")
                
                # Skip if JSON file doesn't exist
                if not os.path.exists(json_path):
                    logger.warning(f"Skipping {table_name}: JSON file not found")
                    import_summary['results'][table_name] = {
                        'status': 'skipped',
                        'reason': 'File not found'
                    }
                    continue
                
                try:
                    # Load data from JSON
                    with open(json_path, 'r') as f:
                        data = json.load(f)
                    
                    logger.info(f"Loaded {len(data)} records from {json_path}")
                    
                    # Import records
                    imported_count = 0
                    errors = []
                    
                    for record_data in data:
                        try:
                            # Create new instance
                            record = model()
                            
                            # Set attributes
                            for key, value in record_data.items():
                                # Handle ISO format datetime strings
                                if isinstance(value, str) and 'T' in value and value.endswith('Z'):
                                    try:
                                        value = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                    except ValueError:
                                        pass
                                
                                setattr(record, key, value)
                            
                            db.session.add(record)
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append(str(e))
                            db.session.rollback()
                    
                    # Commit all records
                    db.session.commit()
                    
                    # Update summary
                    import_summary['results'][table_name] = {
                        'status': 'completed' if not errors else 'partial',
                        'records_imported': imported_count,
                        'errors': errors
                    }
                    
                    logger.info(f"Imported {imported_count} records to {table_name}")
                    
                except Exception as e:
                    logger.error(f"Error importing {table_name}: {e}")
                    db.session.rollback()
                    import_summary['results'][table_name] = {
                        'status': 'error',
                        'error': str(e)
                    }
        
        # Update summary
        import_summary['status'] = 'completed'
        import_summary['completed_timestamp'] = datetime.now().isoformat()
        
        logger.info(f"Import completed with status: {import_summary['status']}")
        return import_summary

    
    def export_to_excel(self, table_name=None):
        """Export data to Excel file(s)"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if table_name:
            # Export single table
            if table_name not in self.models:
                raise ValueError(f"Table not found: {table_name}")
            
            model = self.models[table_name]
            records = model.query.all()
            
            # Convert to list of dictionaries
            data = []
            for record in records:
                record_dict = {}
                for column in model.__table__.columns:
                    value = getattr(record, column.name)
                    # Handle datetime objects
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    record_dict[column.name] = value
                data.append(record_dict)
            
            # Create Excel file in memory
            output = io.BytesIO()
            self._save_as_excel(data, output, table_name)
            output.seek(0)
            
            return output, f"{table_name}_{timestamp}.xlsx"
        else:
            # Export all tables to a single Excel file with multiple sheets
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for table_name in self.dependency_order:
                    if table_name in self.models:
                        model = self.models[table_name]
                        
                        try:
                            # Get all records
                            records = model.query.all()
                            
                            # Convert to list of dictionaries
                            data = []
                            for record in records:
                                record_dict = {}
                                for column in model.__table__.columns:
                                    value = getattr(record, column.name)
                                    # Handle datetime objects
                                    if isinstance(value, datetime):
                                        value = value.isoformat()
                                    record_dict[column.name] = value
                                data.append(record_dict)
                            
                            # Create DataFrame
                            df = pd.DataFrame(data) if data else pd.DataFrame()
                            
                            # Save to sheet
                            sheet_name = table_name[:31]  # Excel limits sheet names to 31 chars
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            # Format sheet
                            worksheet = writer.sheets[sheet_name]
                            
                            # Format header row
                            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                            header_font = Font(color='FFFFFF', bold=True)
                            
                            for cell in worksheet[1]:
                                cell.fill = header_fill
                                cell.font = header_font
                            
                            # Auto-adjust column widths
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                
                                for cell in column:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                
                                adjusted_width = max_length + 2
                                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                            
                        except Exception as e:
                            logger.error(f"Error exporting {table_name} to Excel: {e}")
                            # Create empty sheet with error message
                            df = pd.DataFrame([{"Error": f"Failed to export: {str(e)}"}])
                            sheet_name = table_name[:31]
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            output.seek(0)
            return output, f"database_export_{timestamp}.xlsx"
    
    def get_available_exports(self):
        """Get list of available exports"""
        exports = []
        
        if os.path.exists(self.export_dir):
            # Get all directories in the export directory that look like exports
            export_dirs = [d for d in os.listdir(self.export_dir) 
                          if os.path.isdir(os.path.join(self.export_dir, d)) and 
                          d.startswith('data_exports')]
            
            for export_dir_name in sorted(export_dirs, reverse=True):  # Most recent first
                export_dir = os.path.join(self.export_dir, export_dir_name)
                try:
                    metadata_path = os.path.join(export_dir, 'metadata.json')
                    summary_path = os.path.join(export_dir, 'export_summary.json')
                    
                    export_info = {
                        'path': export_dir,
                        'name': export_dir_name.replace('data_exports_', '').replace('_', ' '),
                        'date': 'Unknown',
                        'records': 0,
                        'size': self._get_directory_size(export_dir),
                        'has_zip': os.path.exists(export_dir + '.zip')
                    }
                    
                    # Load metadata if available
                    if os.path.exists(metadata_path):
                        with open(metadata_path, 'r') as f:
                            metadata = json.load(f)
                            export_info['date'] = datetime.fromisoformat(
                                metadata['export_timestamp']
                            ).strftime('%Y-%m-%d %H:%M')
                    
                    # Load summary if available
                    if os.path.exists(summary_path):
                        with open(summary_path, 'r') as f:
                            summary = json.load(f)
                            export_info['records'] = sum(
                                info.get('records_exported', 0) 
                                for info in summary.values() 
                                if isinstance(info, dict)
                            )
                    
                    exports.append(export_info)
                    
                except Exception as e:
                    logger.error(f"Error processing export {export_dir}: {e}")
        
        return exports
    
    def _get_directory_size(self, path):
        """Get the total size of a directory in bytes, formatted as human-readable string"""
        total_size = 0
        for dirpath, dirnames, filenames in os.walk(path):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                if os.path.exists(filepath):
                    total_size += os.path.getsize(filepath)
        return self._format_file_size(total_size)
    
    def _format_file_size(self, size_bytes):
        """Format file size in bytes as human-readable string"""
        if size_bytes == 0:
            return "0 B"
        
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return f"{s} {size_names[i]}"